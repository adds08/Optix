#!/usr/bin/env python3
"""Generate a seedable structure from TOOL LIST BY NAME.xlsx.

Entity sections: foremen, projects, trailers, small_tools + unmatched_projects.
Rules implemented:
  - TE (primary) + TE-* sheets only; FORMAT and foreman-named sheets ignored.
  - Foremen deduped: normalized key + known-spelling aliases + surname/first-
    initial key (smart matching for hand-typed variants).
  - Projects canonicalized by job # first, then by normalized name; TE-* only
    projects with no TE match land in unmatched_projects.
  - Trailers keyed by TE-xxx; truck always null.
  - Small tools: serial# -> Equipment Department, none -> Purchased Department.
"""
import json
import re
import unicodedata
from collections import OrderedDict
import openpyxl

SRC = "/Users/adds08/Development/Urbaniconstruction/STInventory/docs/data/TOOL LIST BY NAME.xlsx"

def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip()

def norm_key(s):
    return norm(s).lower()

# Hand-typed spelling variants of the same person, observed in the source.
NAME_TOKEN_ALIASES = {
    "jobani": "jovanni",
    "floriberto": "florencio",
}
def smart_name_key(raw):
    n = norm_key(raw)
    tokens = n.split(" ")
    tokens = [NAME_TOKEN_ALIASES.get(t, t) for t in tokens]
    n = " ".join(tokens)
    surname = tokens[-1] if tokens else ""
    initial = tokens[0][0] if tokens and tokens[0] else ""
    return (n, f"{surname}:{initial}")

# Canonical project names by job #, then by fuzzy name.
JOB_CANON = {
    "22018": "Lone Star", "22017": "NEX", "23004": "Colony Phase 12",
    "24003": "Plano Arterial Renewal-2", "22015": "Garland", "24007": "Austin Lane",
    "20011": "DART", "23002": "Richardson", "23010": "Bell", "23009": "Little Elm",
    "24005": "Mesquite",
}
NAME_CANON = {
    "yard": "Equipment Yard", "lonstar": "Lone Star", "lone star": "Lone Star",
    "lonsar": "Lone Star", "lonestar": "Lone Star", "nex": "NEX",
    "colony phase 12": "Colony Phase 12", "colony phase 12 re": "Colony Phase 12",
    "colony": "Colony Phase 12", "plano arterial renewal-2": "Plano Arterial Renewal-2",
    "plano": "Plano Arterial Renewal-2", "garland": "Garland", "austin lane": "Austin Lane",
    "dart": "DART", "richardson": "Richardson", "bell": "Bell", "bell, tx": "Bell",
    "little elm": "Little Elm", "little helm": "Little Elm", "littel elm": "Little Elm",
    "litte elm": "Little Elm", "mesquite": "Mesquite", "traffic control": "Traffic Control",
    "city of kemp": "City of Kemp", "mechanic": "Mechanic", "shop": "Equipment Yard",
}

def canonical_project(name, job):
    if job and job in JOB_CANON:
        return JOB_CANON[job]
    if name:
        return NAME_CANON.get(norm_key(name), name)
    return None

JOB_RE = re.compile(r"#\s*(\d{4,5})")
TRAILER_RE = re.compile(r"^(?:TE[- ]?0*(\d{2,3}))$", re.I)
STATUS_WORDS = {"DAMAGE", "NEED REAPAIRS", "NEED REPAIRS", "STRUCTURE- CONCRETE POURING", "REPAIR", "SHOP"}

def parse_job(cell):
    if cell is None:
        return None
    s = str(cell).strip()
    if s.isdigit() and len(s) >= 4:
        return s
    m = JOB_RE.search(s)
    return m.group(1) if m else None

def split_job(cell):
    """Return (text_before_hash, job) if the cell embeds '#NNNN', else (None, None)."""
    if cell is None:
        return (None, None)
    s = str(cell).strip()
    m = JOB_RE.search(s)
    if not m:
        return (None, None)
    return (s[: m.start()].strip(" #,;:-"), m.group(1))

def is_foreman_word(s):
    s = str(s).strip()
    if not s or TRAILER_RE.match(s) or s.isdigit():
        return False
    if "#" in s:
        return False
    if s.upper() in STATUS_WORDS:
        return False
    return True

wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------------- TE (primary) ----------------
te_ws = wb["TE"]
te_assign = OrderedDict()     # trailer -> {foreman, project, job, status}
te_trailerless = []           # {foreman, project, job}
known_te_foremen = set()

def add_te_project(name, job, trailer=None):
    key = job or (norm_key(name) if name else None)
    if not key:
        return
    canon = canonical_project(name, job)
    entry = te_projects.setdefault(key, {"name": canon or name, "job": job, "trailers": []})
    if canon:
        entry["name"] = canon
    if trailer and trailer not in entry["trailers"]:
        entry["trailers"].append(trailer)

te_projects = OrderedDict()
for r in range(4, te_ws.max_row + 1):
    idv, fv, pv, jv = te_ws.cell(r, 2).value, te_ws.cell(r, 3).value, te_ws.cell(r, 6).value, te_ws.cell(r, 9).value
    if idv is None and fv is None and pv is None and jv is None:
        continue
    trailer = None
    if idv is not None:
        m = TRAILER_RE.match(str(idv).strip())
        if m:
            trailer = f"TE-{int(m.group(1)):03d}"
    job = parse_job(jv) or parse_job(pv)
    proj_name = None
    if pv is not None:
        text, j2 = split_job(pv)
        proj_name = text or norm(pv)
        if j2:
            job = j2
    raw_foreman = norm(fv) if is_foreman_word(fv) else ""
    status = norm(fv).upper() if (fv and not is_foreman_word(fv)) else None
    if trailer:
        # TE repeats some trailer IDs with empty rows further down; never let a
        # blank row overwrite a populated assignment.
        existing = te_assign.get(trailer)
        if existing is None or (existing["foreman"] == "" and raw_foreman):
            te_assign[trailer] = {"foreman": raw_foreman, "project": proj_name, "job": job, "status": status}
        elif existing and not existing["project"] and proj_name:
            existing["project"] = proj_name
            existing["job"] = job
        if raw_foreman:
            known_te_foremen.add(smart_name_key(raw_foreman)[0])
        add_te_project(proj_name, job, trailer)
    else:
        te_trailerless.append({"foreman": raw_foreman, "project": proj_name, "job": job})
        if raw_foreman:
            known_te_foremen.add(smart_name_key(raw_foreman)[0])
        add_te_project(proj_name, job)

# ---------------- TE-* trailer sheets ----------------
trailer_sheets = sorted(
    (n for n in wb.sheetnames if n.upper().startswith("TE-")),
    key=lambda n: int(re.search(r"(\d+)", n).group(1)) if re.search(r"(\d+)", n) else 999,
)

detail = []  # per trailer sheet
for name in trailer_sheets:
    ws = wb[name]
    cells = {(r, c): ws.cell(r, c).value for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None}
    header_row = None
    for (r, c) in cells:
        if str(cells[(r, c)]).strip().upper() in {"DATE", "DESCRIPTION"}:
            header_row = r
            break
    meta_rows = sorted({r for (r, c) in cells if header_row and r < header_row})

    trailer_id, foreman_raw, project_raw, job = None, None, None, None
    for r in meta_rows:
        rowcells = {c: cells[(r, c)] for c in range(1, ws.max_column + 1) if (r, c) in cells}
        id_col = None
        for c, v in rowcells.items():
            if TRAILER_RE.match(str(v).strip()):
                trailer_id = f"TE-{int(TRAILER_RE.match(str(v).strip()).group(1)):03d}"
                id_col = c
                break
        if id_col:
            for c in range(id_col + 1, ws.max_column + 1):
                if c in rowcells and is_foreman_word(rowcells[c]):
                    foreman_raw = norm(rowcells[c])
                    break
        if foreman_raw is None and trailer_id is None:
            for c, v in rowcells.items():
                if is_foreman_word(v):
                    foreman_raw = norm(v)
                    break
        for c, v in rowcells.items():
            if c == id_col:
                continue
            text, j2 = split_job(v)
            if j2:
                job = j2
                # combined "Name #job"?  A person's name before '#' means the
                # foreman and job share one cell (Te-035).
                if text and foreman_raw is None and smart_name_key(text)[0] in known_te_foremen:
                    foreman_raw = norm(text)
                elif text and project_raw is None:
                    project_raw = text
                continue
            s = str(v).strip()
            if s.isdigit() and len(s) >= 4 and not job:
                job = s
            elif project_raw is None and is_foreman_word(v) is False and s and not TRAILER_RE.match(s):
                # plain project-name-looking cell (no #): only if not a name
                if smart_name_key(s)[0] not in known_te_foremen:
                    project_raw = norm(v)

    # tools
    hdr = {str(ws.cell(header_row, c).value).strip().upper(): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
    def cell(r, lbl):
        c = hdr.get(lbl)
        return ws.cell(r, c).value if c else None
    tools = []
    for r in range((header_row or 1) + 1, ws.max_row + 1):
        desc, make = cell(r, "DESCRIPTION"), cell(r, "MAKE")
        model, serial, other = cell(r, "MODEL"), cell(r, "SERIAL #"), cell(r, "OTHER")
        qtyv = cell(r, "QTY")
        if all(v is None for v in (desc, make, model, serial)) and (qtyv is None or str(qtyv).strip() == ""):
            continue
        if desc is None and (qtyv is None or str(qtyv).strip() in ("", "1")) and serial is None and make is None:
            continue
        if desc is None and serial is not None and (qtyv is None or str(qtyv).strip() in ("", "1")) and make is None and model is None:
            if tools:
                tools[-1]["serials"].append(str(serial).strip())
            continue
        tools.append({
            "qty": int(qtyv) if str(qtyv or "").strip().isdigit() else 1,
            "description": norm(desc) or None,
            "make": norm(make) or None,
            "model": norm(model) or None,
            "serials": [str(serial).strip()] if serial is not None else [],
            "other": norm(other) or None,
        })
    detail.append({
        "sheet": name, "trailer": trailer_id, "foreman": foreman_raw,
        "project": project_raw, "job": job, "tools": tools,
    })

# ---------------- Foremen dedup ----------------
foremen = []          # {id, external_id, name, key}
foremen_by_key = {}
def get_foreman(raw):
    if not raw:
        return None
    exact, fuzzy = smart_name_key(raw)
    if exact in foremen_by_key:
        return foremen_by_key[exact]
    # smart merge: same surname + first initial (catches Jobani/Jovanni etc.)
    for f in foremen:
        if f["fuzzy"] == fuzzy:
            # prefer the display name already stored if near-identical length
            if abs(len(f["name"]) - len(norm(raw))) <= 3:
                f["aliases"].append(norm(raw))
                foremen_by_key[exact] = f["id"]
                return f["id"]
    fid = len(foremen) + 1
    entry = {"id": fid, "external_id": f"FM-{fid:03d}", "name": norm(raw), "fuzzy": fuzzy, "aliases": []}
    foremen.append(entry)
    foremen_by_key[exact] = fid
    return fid

for a in te_assign.values():
    if a["foreman"]:
        get_foreman(a["foreman"])
for f in te_trailerless:
    if f["foreman"]:
        get_foreman(f["foreman"])
for d in detail:
    if d["foreman"]:
        get_foreman(d["foreman"])

# ---------------- Projects + reconciliation ----------------
def project_record(name, job):
    canon = canonical_project(name, job)
    key = job or norm_key(canon or name)
    if key in te_projects:
        return ("matched", te_projects[key]["name"], te_projects[key]["job"])
    # a project keyed only by name in TE?
    if canon:
        for k, p in te_projects.items():
            if norm_key(p["name"]) == norm_key(canon):
                return ("matched", p["name"], p["job"])
    return ("unmatched", canon or name, job)

unmatched_projects = []
for d in detail:
    if d["project"] or d["job"]:
        status, pname, pjob = project_record(d["project"], d["job"])
        d["_project_status"] = status
        d["_project_name"] = pname
        d["_project_job"] = pjob
        if status == "unmatched":
            unmatched_projects.append({
                "sheet": d["sheet"], "trailer_id": d["trailer"],
                "project_name_in_sheet": d["project"], "job_in_sheet": d["job"],
                "canonical_name": pname, "canonical_job": pjob,
            })

# ---------------- Reconciliation & conflicts ----------------
from difflib import SequenceMatcher

conflicts = {
    "unmatched_projects": unmatched_projects,
    "foreman_mismatches": [],
    "project_mismatches": [],
    "trailer_anomalies": [],
    "merged_name_variants": [],
    "unresolved_name_variants": [],
    "notes": [],
}

def same_person(a, b):
    """True when the dedup logic would merge these two names into one person:
    identical normalized keys, or same surname + first initial with names close
    enough in length that the difference is a spelling variant."""
    ea, fa = smart_name_key(a)
    eb, fb = smart_name_key(b)
    if ea == eb:
        return True
    return fa == fb and abs(len(norm(a)) - len(norm(b))) <= 3

# --- foreman mismatches: TE (primary) vs the trailer's detail sheet ---
for d in detail:
    if not d["trailer"] or d["trailer"] not in te_assign:
        continue
    te_f = te_assign[d["trailer"]].get("foreman") or ""
    sh_f = d["foreman"] or ""
    if not te_f or not sh_f:
        continue
    if norm_key(te_f) == norm_key(sh_f):
        continue  # same name, different case/spacing - not a conflict
    if same_person(te_f, sh_f):
        conflicts["merged_name_variants"].append({
            "trailer_id": d["trailer"], "sheet": d["sheet"],
            "te_name": te_f, "sheet_name": sh_f,
            "severity": "info",
            "resolution": "merged (spelling variant)",
        })
    else:
        conflicts["foreman_mismatches"].append({
            "trailer_id": d["trailer"], "sheet": d["sheet"],
            "te_name": te_f, "sheet_name": sh_f,
            "severity": "conflict",
            "resolution": "unresolved - TE primary used; verify which is current",
        })

# --- project mismatches: TE project/job vs the detail sheet's project/job ---
for d in detail:
    if not d["trailer"] or d["trailer"] not in te_assign:
        continue
    a = te_assign[d["trailer"]]
    te_job, sh_job = a.get("job"), d.get("job")
    te_proj, sh_proj = a.get("project"), d.get("project")
    if (te_job or sh_job) and te_job != sh_job:
        conflicts["project_mismatches"].append({
            "trailer_id": d["trailer"], "sheet": d["sheet"],
            "te_project": te_proj, "te_job": te_job,
            "sheet_project": sh_proj, "sheet_job": sh_job,
            "severity": "conflict",
            "resolution": "unresolved - TE primary used; different job numbers",
        })
    elif (te_job or sh_job) and te_job == sh_job and norm_key(te_proj or "") != norm_key(sh_proj or ""):
        conflicts["project_mismatches"].append({
            "trailer_id": d["trailer"], "sheet": d["sheet"],
            "te_project": te_proj, "te_job": te_job,
            "sheet_project": sh_proj, "sheet_job": sh_job,
            "severity": "info",
            "resolution": "same job number, different spelling - merged by job",
        })

# --- trailer anomalies ---
for tid, a in te_assign.items():
    sheet = next((d for d in detail if d["trailer"] == tid), None)
    if not a["foreman"]:
        if a["status"]:
            conflicts["trailer_anomalies"].append({
                "trailer_id": tid, "kind": "no_foreman", "te_foreman": a["foreman"],
                "te_status": a["status"], "sheet_foreman": sheet["foreman"] if sheet else None,
                "note": "trailer flagged in source (damage/repair/use note); no foreman assigned",
            })
        elif not sheet:
            conflicts["trailer_anomalies"].append({
                "trailer_id": tid, "kind": "empty_entry", "te_foreman": None,
                "note": "trailer id present in TE but no foreman, no status, no detail sheet",
            })
        else:
            conflicts["trailer_anomalies"].append({
                "trailer_id": tid, "kind": "sheet_only", "te_foreman": None,
                "sheet_foreman": sheet["foreman"], "sheet_project": sheet.get("_project_name"),
                "sheet_job": sheet.get("_project_job"),
                "note": "no TE row; trailer + foreman taken from TE-* detail sheet",
            })
    if sheet and not (a["job"] or a["project"]) and not (sheet.get("_project_job") or sheet.get("_project_name")):
        conflicts["trailer_anomalies"].append({
            "trailer_id": tid, "kind": "no_project", "te_foreman": a["foreman"],
            "sheet_foreman": sheet["foreman"] if sheet else None,
            "note": "no project or job number anywhere for this trailer",
        })

# trailers in detail sheets that never appear in TE at all
detail_ids = {d["trailer"] for d in detail if d["trailer"]}
for d in detail:
    if d["trailer"] and d["trailer"] not in te_assign and d["foreman"]:
        conflicts["trailer_anomalies"].append({
            "trailer_id": d["trailer"], "kind": "no_te_row", "te_foreman": None,
            "sheet_foreman": d["foreman"], "sheet_project": d.get("_project_name"),
            "sheet_job": d.get("_project_job"),
            "note": "trailer only exists in its TE-* detail sheet",
        })

# --- unresolved name variants: same first name, similar surname, not merged ---
seen = []
for i, f in enumerate(foremen):
    for g in foremen[i + 1:]:
        ft = norm(f["name"]).split()
        gt = norm(g["name"]).split()
        if not ft or not gt:
            continue
        if ft[0].lower() == gt[0].lower():
            fsur, gsur = ft[-1], gt[-1]
            ratio = SequenceMatcher(None, fsur.lower(), gsur.lower()).ratio()
            if 0.4 <= ratio < 1.0:
                conflicts["unresolved_name_variants"].append({
                    "name_a": f["name"], "id_a": f["external_id"],
                    "name_b": g["name"], "id_b": g["external_id"],
                    "surname_similarity": round(ratio, 3),
                    "note": "same first name and near-matching surname - likely the same person, verify before merging",
                })

conflicts["notes"].append(
    "TE primary sheet is authoritative for trailer->foreman->project; TE-* detail sheets"
    " fill gaps and their tools attach to the sheet foreman."
)
conflicts["notes"].append(
    "Duplicated TE rows (blank repeats of TE-029/030/031/033) were discarded; a populated"
    " row was never overwritten by a blank one."
)
# TE is primary, but a trailer may only exist in its TE-* detail sheet (e.g.
# TE-032, TE-034 have no foreman row in TE). The sheet fills the gap.
trailers = []
seen_trailer_ids = set()
for tid, a in te_assign.items():
    sheet = next((d for d in detail if d["trailer"] == tid), None)
    foreman = a["foreman"] or (sheet["foreman"] if sheet else None)
    # resolve the project (canonical name + unique id) early
    pname = a["project"] or (sheet.get("_project_name") if sheet else None)
    pjob = a["job"] or (sheet.get("_project_job") if sheet else None)
    status, pname_c, pjob_c = project_record(pname, pjob) if (pname or pjob) else ("none", None, None)
    pid = None
    if pname_c or pjob_c:
        pid = f"proj-{norm_key(pname_c or 'unknown').replace(' ', '-')}" + (f"-{pjob_c}" if pjob_c else "")
    if not foreman and a["status"]:
        trailers.append({
            "id": tid, "foreman_id": None, "foreman_name": None,
            "project_id": pid, "project_name": pname_c, "job": pjob_c,
            "truck": None, "status": a["status"].lower().replace(" ", "_"),
            "note": "no foreman assigned in source",
        })
        seen_trailer_ids.add(tid)
        continue
    if not foreman:
        continue
    fid = get_foreman(foreman)
    trailers.append({
        "id": tid, "foreman_id": fid, "foreman_name": norm(foreman),
        "project_id": pid, "project_name": pname_c, "job": pjob_c,
        "truck": None, "status": "active",
        "note": "assigned in TE-* detail sheet only" if not a["foreman"] else None,
    })
    seen_trailer_ids.add(tid)

# Trailers that only live in a detail sheet (no TE row at all).
for d in detail:
    if d["trailer"] and d["trailer"] not in seen_trailer_ids and d["foreman"]:
        fid = get_foreman(d["foreman"])
        status, pname2, pjob2 = project_record(d.get("_project_name"), d.get("_project_job"))
        pid = f"proj-{norm_key(pname2 or 'unknown').replace(' ', '-')}" + (f"-{pjob2}" if pjob2 else "") if pname2 or pjob2 else None
        trailers.append({
            "id": d["trailer"], "foreman_id": fid, "foreman_name": norm(d["foreman"]),
            "project_id": pid, "project_name": pname2, "job": pjob2,
            "truck": None, "status": "active",
            "note": "trailer appears in TE-* sheet only",
        })

# ---------------- Small tools ----------------
small_tools = []
seq = 0
for d in detail:
    tid = d["trailer"]
    foreman_id = get_foreman(d["foreman"]) if d["foreman"] else None
    pname, pjob = d.get("_project_name"), d.get("_project_job")
    for t in d["tools"]:
        for i, serial in enumerate(t["serials"] or [None]):
            seq += 1
            dept = "equipment" if serial else "purchased"
            small_tools.append({
                "id": f"TOOL-{seq:04d}",
                "qty": t["qty"] if not serial else 1,
                "description": t["description"],
                "make": t["make"],
                "model": t["model"],
                "serial": serial,
                "other": t["other"],
                "department": dept,
                "trailer_id": tid,
                "foreman_id": foreman_id,
                "project_name": pname,
                "job": pjob,
            })

# ---------------- Output ----------------
output = {
    "meta": {
        "source_file": "docs/data/TOOL LIST BY NAME.xlsx",
        "source_sheets_processed": ["TE"] + trailer_sheets,
        "sheets_ignored": [n for n in wb.sheetnames if n not in ["TE"] + trailer_sheets],
        "generated_by": "docs/data/generate_seed.py",
        "notes": [
            "Trailer 'truck' left null per spec.",
            "Small tool department: serial present -> Equipment, absent -> Purchased.",
            "Unmatched projects (if any) exported to unmatched_projects.json for review.",
            "TE-021 sheet foreman 'Jose Danery' differs from TE primary 'Gabriel Villareal' - kept both, trailer uses TE primary.",
            "TE-023 sheet 'RAFAEL JIMENEZ' vs TE primary 'RAFAEL JAIMES' - kept separate (different surnames).",
        ],
    },
    "foremen": [{k: f[k] for k in ("id", "external_id", "name", "aliases")} for f in foremen],
    "projects": [
        {"id": f"proj-{norm_key(p['name']).replace(' ', '-')}" + (f"-{p['job']}" if p["job"] else ""), "externalId": p["job"], "name": p["name"], "job": p["job"], "trailer_ids": p["trailers"]}
        for p in te_projects.values()
    ],
    "trailers": trailers,
    "small_tools": small_tools,
    "unmatched_projects": unmatched_projects,
}

out_dir = "/Users/adds08/Development/Urbaniconstruction/STInventory/docs/data"
with open(f"{out_dir}/seed_from_tools_list.json", "w") as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

# Full reconciliation report: everything a human should review before loading
# the seed - unmatched projects, foreman/project conflicts between TE and the
# TE-* sheets, trailer anomalies, and name variants (merged and unresolved).
report = {
    "purpose": "Human review queue for docs/data/seed_from_tools_list.json. "
               "Nothing here is auto-resolved silently - every entry needs a "
               "human decision before the seed is trusted.",
    "summary": {
        "unmatched_projects": len(conflicts["unmatched_projects"]),
        "foreman_mismatches": len(conflicts["foreman_mismatches"]),
        "project_mismatches": len(conflicts["project_mismatches"]),
        "trailer_anomalies": len(conflicts["trailer_anomalies"]),
        "merged_name_variants": len(conflicts["merged_name_variants"]),
        "unresolved_name_variants": len(conflicts["unresolved_name_variants"]),
    },
    "unmatched_projects": conflicts["unmatched_projects"],
    "foreman_mismatches": conflicts["foreman_mismatches"],
    "project_mismatches": conflicts["project_mismatches"],
    "trailer_anomalies": conflicts["trailer_anomalies"],
    "merged_name_variants": conflicts["merged_name_variants"],
    "unresolved_name_variants": conflicts["unresolved_name_variants"],
    "notes": conflicts["notes"],
}
import datetime
report["generated_at"] = datetime.datetime.now(datetime.timezone.utc).isoformat()

with open(f"{out_dir}/reconciliation_report.json", "w") as f:
    json.dump(report, f, indent=2, ensure_ascii=False)

print("\n=== RECONCILIATION REPORT SUMMARY ===")
for k, v in report["summary"].items():
    print(f"   {k}: {v}")

# Diagnostics
print(f"foremen: {len(foremen)}")
for f in foremen:
    print(f"   {f['id']:3d} {f['external_id']} {f['name']!r} aliases={f['aliases']}")
print(f"\nprojects: {len(te_projects)}")
for p in te_projects.values():
    print(f"   {p['name']!r} job={p['job']!r} trailers={p['trailers']}")
print(f"\ntrailers: {len(trailers)}")
print(f"small_tools: {len(small_tools)}  (equipment={sum(1 for t in small_tools if t['department']=='equipment')}, purchased={sum(1 for t in small_tools if t['department']=='purchased')})")
print(f"unmatched_projects: {len(unmatched_projects)}")
for u in unmatched_projects:
    print(f"   {u}")
