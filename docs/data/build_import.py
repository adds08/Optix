#!/usr/bin/env python3
"""Build import templates + a rejects queue from the two real-data sources.

Sources
  A. docs/data/TOOL LIST BY NAME NEW.xlsx
       'TE'      — ENCLOSED TRAILER ASSIGNED LIST: trailer -> foreman -> project.
                   Authoritative for who holds which trailer.
       'TE-*'    — one sheet per trailer: its own foreman/project header row plus
                   the small tools aboard it.
       'FORMAT'  — a blank template. Ignored.
       The person-named sheets present in the OLD workbook (mechanics) are gone
       from this one, which is what was asked for; nothing here reads them.
  B. The company-truck PDF, transcribed into TRUCKS below.

Nothing is written to any database. Two kinds of output:
  import/*.csv    rows clean enough to load
  import/rejects.{json,csv}  every row that is NOT safe to load, with the reason

The reject rules exist because the target schema has real uniqueness and
not-null constraints (vehicle.unit NOT NULL, asset.tag, employee identity) and a
load that guesses a missing id produces a collision or a second custodian, which
is the one failure mode this product cannot absorb.
"""
import csv
import json
import re
import unicodedata
from collections import OrderedDict, defaultdict

import openpyxl

ROOT = "/Users/adds08/Development/Urbaniconstruction/STInventory"
SRC = f"{ROOT}/docs/data/TOOL LIST BY NAME NEW.xlsx"
OUT = f"{ROOT}/docs/data/import"

# The yard is NOT a job, so it carries no job number. Job 24002 is a separate
# real job (it has its own superintendent and surveyor in the truck list) that
# somebody also typed against the yard on TE-007 — see DECISIONS.
#
# The name must stay exactly "Equipment Yard": apps/web/app/(app)/jobsites/page.tsx
# matches the yard by name (`YARD_PROJECT_NAME = "equipment yard"`), so calling
# this project "YARD" would make isYardProject() miss it and draw the yard as an
# ordinary job card — the bug that page exists to fix.
YARD = "Equipment Yard"

# ---------------------------------------------------------------- DECISIONS
# Human rulings on the identity collisions, 2026-09-01. Recorded here rather
# than applied silently, because every one of them decides whether two rows are
# one person — and a wrong merge produces a custody record for somebody who does
# not exist. Each entry says who ruled and on what evidence.
#
# 1. Juan Martinez: TWO people, not three.
#    "JUAN MARTINEZ" (TE + TE-028 sheet, NEX #22017) and "Juan Martinez (1975)"
#    (TRK-044, job 22017) are the same man — same job, same role. The "(1975)"
#    was added by somebody at Urban precisely because a second Juan Martinez
#    exists, which is evidence FOR the split, not against it.
#    "Juan Carlos Martinez" (TRK-032, job 25001) is that second man: different
#    job, different crew. Kept separate.
#
# 2. The other five near-matches are KEPT SEPARATE, deliberately.
#    LOZA SR., Abarca, Medina, Almaguer and Capuchino each look like one person
#    spelled two ways, and an earlier pass at this data (generate_seed.py's
#    NAME_TOKEN_ALIASES) merged two of them. Urban's call is to split: a wrong
#    merge silently hands one man another man's tools, while a wrong split leaves
#    two obvious duplicates a human can merge later from the register. The
#    reversible error is the one to make.
#
# 3. Equipment Yard is the yard itself and carries NO job number.
#    Job 24002 becomes its own real project — it has a superintendent (TRK-036),
#    a surveyor (TRK-038), a traffic control foreman (TRK-020) and a field
#    engineer (TRK-045), which is a job crew, not yard staff. The duplicate
#    "Equipment Yard / 24002" project row is what draws a third yard card today.
#
# 4. TE-027 has NO foreman for now.
#    FELIPE PORTILLO appeared as custodian of both TE-017 and TE-027 (both DART
#    #20011), and one person holding two enclosed trailers is not how Urban runs.
#    Rather than invent a second Felipe, TE-027 is loaded with no custodian and
#    its tools stay with the TRAILER — located on TE-027, assigned to nobody.
#    Whoever actually runs TE-027 can be set in the app in one action.
MERGE_PEOPLE = {
    "juan martinez": "Juan Martinez (1975)",
    "juan martinez (1975)": "Juan Martinez (1975)",
}
# Pairs a human looked at and chose NOT to merge. Recorded so the next run does
# not re-raise them as open questions.
KEPT_SEPARATE = [
    ("FLORENCIO LOZA SR.", "FLORIBERTO LOZA SR."),
    ("Jobani Abarca", "JOVANI ABARCA"),
    ("Gilmer Medina", "Gilmar Medina"),
    ("Romualdo", "Romualdo Almaguer"),
    ("Alejandro Capuchino", "Alejandro Aranda Capuchino"),
]
# Trailers deliberately loaded with no custodian (decision 4).
NO_CUSTODIAN_TRAILERS = {"TE-027"}

rejects = []


def reject(entity, ident, reason, severity, **raw):
    rejects.append({
        "entity": entity, "id": ident, "reason": reason,
        "severity": severity, "raw": {k: v for k, v in raw.items() if v is not None},
    })


def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip()


def nkey(s):
    return norm(s).lower()


def pkey(s):
    """The key a name is stored under in the people registry, after any
    human-ruled merge. nkey() alone is wrong for a merged spelling."""
    n = norm(s)
    return nkey(MERGE_PEOPLE.get(nkey(n), n))


# ---------------------------------------------------------------- people
# A name is the single easiest thing to get wrong here: the sources are
# hand-typed, and two different people genuinely share a surname. So this
# NEVER merges on a fuzzy match -- it merges only on an exact normalized
# name, and REPORTS anything that looks close enough to be the same person.
class People:
    def __init__(self):
        self.by_name = OrderedDict()   # nkey -> record
        self.sources = defaultdict(list)

    def add(self, raw, source, position=None, job=None):
        n = norm(raw)
        if not n:
            return None
        # Only an explicit, human-ruled merge collapses two spellings (DECISIONS 1).
        n = MERGE_PEOPLE.get(nkey(n), n)
        k = nkey(n)
        if k not in self.by_name:
            self.by_name[k] = {
                "name": n, "positions": [], "jobs": [], "sources": [],
            }
        rec = self.by_name[k]
        if position and position not in rec["positions"]:
            rec["positions"].append(position)
        if job and job not in rec["jobs"]:
            rec["jobs"].append(job)
        if source not in rec["sources"]:
            rec["sources"].append(source)
        return k

    def ambiguities(self):
        """Pairs a human must rule on before these become employee rows."""
        out = []
        keys = list(self.by_name)
        for i, a in enumerate(keys):
            for b in keys[i + 1:]:
                ta, tb = a.split(), b.split()
                if not ta or not tb:
                    continue
                # same surname, same first initial, different full name
                if ta[-1] == tb[-1] and ta[0][0] == tb[0][0] and a != b:
                    out.append((self.by_name[a], self.by_name[b], "same surname + first initial"))
                # one name is a strict subset of the other's tokens
                elif set(ta) < set(tb) or set(tb) < set(ta):
                    out.append((self.by_name[a], self.by_name[b], "one name is a subset of the other"))
        return out


people = People()


def canon_name(raw):
    """The canonical display name for a person, after any human-ruled merge.
    Every CSV that names a custodian must use this, or a merged spelling leaves
    a row pointing at a person the employee list does not contain."""
    if not raw:
        return None
    k = pkey(raw)
    return people.by_name[k]["name"] if k in people.by_name else norm(raw)



# ---------------------------------------------------------------- trucks (PDF)
# Transcribed from "Latest update on company truck.pdf". Kept verbatim,
# misspellings included -- correcting a name here would silently invent a
# person. year is left as the source had it.
# unit, holder, position, job, model, vin, license, toll, year
TRUCKS = [
    ("SUV-001", "Anup Tamrakar", "CEO", None, "ROGUE", "5N1BT3AA2PC906289", "TWV0162", None, 2023),
    ("SUV-2", "Equipment Department", "ASstant manager", "Estimate department", "4RUNNER", "JTEZU5JR8B5029107", "CX9-P896", None, 2011),
    ("TRK-003", "Francisco Perez", "Field Mechanic", "10001", "RAM 3500", "3C63RRGLXJG305453", "LFL1796", "Yes", 2018),
    ("TRK-012", "Romualdo Almaguer", "Traffic Control Foreman", "26002", "TUNDRA", "5TFRM5F10KX143789", "RYB8946", None, 2019),
    ("TRK-013", "Gabriel Garcia", "Forman (Pavement team)", "25011/25012", "F-350", "1FD8W3HT4KEE60507", "NCCJ-2905", "yes", 2019),
    ("TRK-014", "Herman Garza", "Formn", "25011", "F-150", "1FTEW1EP7LKE86086", "SBG-7670", "Yes", 2020),
    ("TRK-015", "Christhopher John", "Mechanic", "10001", "F-550", "1FDUF5GT6FEC83914", "1N17485", "Yes", 2015),
    ("TRK-016", "Equipment Department( out of service)remove it from portal", "CDL", "New yard", "RAM-3500", "3C63RRHL4MG616766", "THR-6180", None, 2021),
    ("TRK-017", "Prajwal Shrsetha", "Equipment Manager", "10001", "Silverado 1500", "2GCUDDEDXP1122128", "THN-9189", "Yes", 2023),
    ("TRK-018", "Raimando Fabela", "CDL", "10001", "RAM-3500", "3C63RRHL0NG188602", "TCR-2073", "Yes", 2022),
    ("TRK-019", "Arnulfo Galvan", "CDL", "22017", "F550", "1FDUF5GT6GEA17441", "TTT7348", "Yes", 2016),
    ("TRK-020", "Gilmar Medina", "Traffic Control Foreman", "24002", "F-250", "1FT7W2B63LEC90333", "VHX7593", "Yes", 2020),
    ("TRK-021", "Luis Luna", "Superintendent", "25001", "F-150", "1FTEW1CP0PKE81997", "XLW1035", "Yes", 2023),
    ("TRK-022", "Dago Ruiz", "Saftey Manager", "Safety Department", "F-150", "1FTEW1CP6PKE66565", "TNW8872", "Yes", 2023),
    ("TRK-024", "Yajju Maharjan", "Assitant Project Manager,24005", "24005/23002/24006", "F-150", "1FTEW1CP8PFD31984", "TTY0037", "Yes", 2023),
    ("TRK-025", "Anup Tamrakar", "CEO", None, "F-150", "1FTEW1CPXPKF37377", "TTY0036", "Yes", 2023),
    ("TRK-026", "Richard Willis", "Project Manager", "25001/24014", "F-150", "1FTEW1CP4PFD31495", "TTY0035", "Yes", 2023),
    ("TRK-027", "Jovani Abarca", "Forman", "25015", "F-250", "1FT7W2AA8REC08644", "TYN0462", "No", 2024),
    ("TRK-028", "Joel Gonzalez", "Field Engineer", "25001/24014", "F-150", "1FTEW1CP4PKF88681", "TVZ4180", "No", 2023),
    ("TRK-029", "Aryana Jimenez", "Field Engineer", "22017", "F-150", "1FTEW1C55PKF74349", "TVZ4179", None, 2023),
    ("TRK-030", "Julius Guerra", "Superintendent", "26002", "F-150", "1FTEW1CPXPKG21764", "TVZ3848", None, 2023),
    ("TRK-031", "Ismael Davilia", "General Superintendent( All jobs bridge)", "22017", "2500", "1GC4YREY3MF271120", "SZL-7694", None, 2021),
    ("TRK-032", "Juan Carlos Martinez", "Forman", "25001", "F-250", "1FT7W2AA5REC45313", "VGC1264", None, 2024),
    ("TRK-033", "Shane Thiesse", "Field Engineer", "22018", "F-150", "1FTEW1KP6RKD61812", "TWD3946", None, 2024),
    ("TRK-034", "Mohamad Nooh Ansari", "Project Manager", "24014/24019", "F-150", "1FTEW1KP6RKD65688", "TWD3850", None, 2024),
    ("TRK-035", "Carlos Espinosa", "Assitant Manager", "10001", "F-150", "1FTEW1KP6RKD63997", "TWD3850", None, 2024),
    ("TRK-036", "Alex Garcia", "Superintendent", "24002", "F-150", "1FTEW1KP6RKD65679", "TWD3774", "Yes", 2024),
    ("TRK-037", "Saul Maldonado", "Project Engineer,", "25008", "F-150", "1FTEW1KP6RKD64963", "TWD3945", None, 2024),
    ("TRK-038", "Orlando Gurera", "Surveyor", "24002", "F-150", "1FTEW1CB5PKF73157", "VPL7462", None, 2023),
    ("TRK-039", "John Solorzano", "PM", "2,201,826,002", "F-150", "1FTEW1CB2PKF73715", "VPL7461", None, 2023),
    ("TRK-040", "Daniel Stewart", "Quality Manager", "25008", "F-150", "1FTEW1CB5PKF73191", "VPM2222", None, 2023),
    ("TRK-041", "Fernando Aguillon", "Forman", "26007", "F-250", "1FT7W2BA2RED13078", "VNX8590", None, 2024),
    ("TRK-042", "Zelvin Perez", "Forman", "22018", "F-250", "1FT7W2BA2RED13873", "VJB0953", None, 2024),
    ("TRK-043", "Bryon Arevalo", "Forman", "24015", "F-250", "1FT7W2BA2RED13217", "VMT5561", None, 2024),
    ("TRK-044", "Juan Martinez (1975)", "Forman", "22017", "F-250", "1FT7W2BA4REE58008", "VSN3520", None, 2024),
    ("TRK-045", "Kevin Zalasar", "field Engineer", "24002", "F-150", "1FTEW1CB8PKF73881", "VPM2223", None, 2023),
    ("TRK-046", "Ruben Palacio", "Super,22017", "26002", "F-150", "1FTEW1CB9PKF72478", "VPM2224", None, 2023),
    ("TRK-047", "Alejandro Aranda Capuchino", "Forman", "22018", "F-250", "1FT7W2BA4REF10222", "vsn3618", None, 2024),
    ("TRK-048", "Raymandu Rodriguiz", "Superintendent", "22018", "F-250", "1FT7W2BA9REF11446", "vsn3619", None, 2024),
    ("TRK-049", "Santiago Gonzalez Franco", "Field Engineer", "21026/26002", "F-150", "1FTEW1KP3RKE78712", "WGN6711", "yes", 2024),
    ("TRK-050", "Christain Carroll", "Field Engineer", "22018", "F-150", "1FTEW1KP6SKD12495", "WGN6712", "yes", 2024),
    ("TRK-051", "Eduardo Reyana", "Project Engineer", "25001 & 25002", "F-150", "1FTEW1KP9SKD12118", "WGN6718", "yes", 2024),
    ("TRK-052", "Joel de Hoyos", "Forman", "22017", "F-250", "1FT7W2BA7REF36099", "WGN6713", "yes", 2024),
    ("TRK-053", "Rafael Herrera", "Superintendent", "24014", "F-250", "1FT7W2BA3REF23205", "WDS3074", "yes", 2024),
    ("TRK-054", "Yunior Ruiz", "Forman", "22018", "F-250", "1FT7W2BA9REF22950", "WGN6717", "yes", 2024),
    ("TRK-055", "Misael Hernandez Rivas", "Forman", "26002", "F-250", "1FT7W2BA7REF23322", "WJS5646", "Yes", 2024),
    ("TRK-056", "Gabrial Zuniga Franco", "Forman", "24014", "F-250", "1FT7W2BA8REF23233", "WDS2705", "Yes", 2024),
    ("TRK-057", "Armando Morado", "Paving Mechanic", "25011", "F-650", "1FDFF6LT1TDA09011", "YDT9949", None, 2025),
    ("TRK-058", "Romualdo Almaguer", "Traffic control", "26002", "F-250", "1FD7X2BA7TEC24591", "Temproary", None, 2026),
    ("TRK-10001", "Jose Zamora", "Mechanic", "10001", "F550", "1FDFF6L2RDA34315", "XCY9662", "Yes", 2024),
]

# A holder cell that is not a person. These trucks are real; their custodian is not.
NOT_A_PERSON = re.compile(r"department|out of service|remove it from portal", re.I)
# A job cell that is not a job number.
JOB_NUM = re.compile(r"^\d{4,5}$")

VALID_VIN = re.compile(r"^[A-HJ-NPR-Z0-9]{17}$")


def clean_job(cell):
    """Return (job, problem). A job is a single 4-5 digit number or nothing."""
    if cell is None:
        return (None, None)
    s = norm(cell)
    if JOB_NUM.match(s):
        return (s, None)
    if re.search(r"[,/&]", s):
        return (None, f"multiple or mangled job numbers: {s!r}")
    return (None, f"not a job number: {s!r}")


# ---------------------------------------------------------------- read TE
wb = openpyxl.load_workbook(SRC, data_only=True)
te = wb["TE"]

TRAILER_RE = re.compile(r"^TE[- ]?0*(\d{1,3})$", re.I)
TRUCK_RE = re.compile(r"^TRK[- ]?0*(\d{1,5})$", re.I)
# A status word sitting where a foreman's name belongs: the trailer exists but
# nobody holds it.
STATUS_WORDS = re.compile(r"damage|repair|reapair|concrete pouring|shop", re.I)

JOB_CANON = {
    "22018": "Lone Star", "22017": "NEX", "23004": "Colony Phase 12",
    "24003": "Plano Arterial Renewal-2", "22015": "Garland", "24007": "Austin Lane",
    "20011": "DART", "23002": "Richardson", "23010": "Bell", "23009": "Little Elm",
    "24005": "Mesquite",
}
NAME_CANON = {
    "yard": YARD, "shop": YARD,
    "lonstar": "Lone Star", "lonsar": "Lone Star", "lonestar": "Lone Star",
    "lone star": "Lone Star", "nex": "NEX", "colony phase 12": "Colony Phase 12",
    "colony phase 12 re": "Colony Phase 12",
    "plano arterial renewal-2": "Plano Arterial Renewal-2", "garland": "Garland",
    "austin lane": "Austin Lane", "dart": "DART", "richardson": "Richardson",
    "bell": "Bell", "little elm": "Little Elm", "litte elm": "Little Elm",
    "mesquite": "Mesquite", "traffic control": "Traffic Control",
    "city of kemp": "City of Kemp", "mechanic": "Mechanic",
}


def canon_project(name, job):
    if job and job in JOB_CANON:
        return JOB_CANON[job]
    if name:
        return NAME_CANON.get(nkey(name), norm(name))
    return None


trailers = OrderedDict()        # unit -> record
te_seen_blank_repeat = []

for r in range(4, te.max_row + 1):
    idv = te.cell(r, 2).value
    holder = te.cell(r, 3).value
    proj = te.cell(r, 6).value
    job = te.cell(r, 9).value
    if all(v is None for v in (idv, holder, proj, job)):
        continue

    ids = norm(idv)
    mt = TRAILER_RE.match(ids) if ids else None
    mk = TRUCK_RE.match(ids) if ids else None

    jobnum, jobproblem = clean_job(job)
    pname = canon_project(proj, jobnum)

    # A truck in the trailer list. Real, but it is not a trailer.
    if mk:
        reject("vehicle", ids, "a truck row inside the trailer list (TE sheet); "
               "cross-check against the truck source before loading",
               "warn", holder=norm(holder), project=norm(proj), sheet="TE", row=r)
        if holder and not NOT_A_PERSON.search(norm(holder)):
            people.add(holder, "TE", job=jobnum)
        continue

    if not mt:
        # rows 51+ : a foreman and a project, but no trailer id at all
        if holder:
            hn = norm(holder)
            if STATUS_WORDS.search(hn):
                continue
            # 'SAN ANTONIO' sits in the foreman column but is a place
            if hn.isupper() and len(hn.split()) == 2 and hn in ("SAN ANTONIO",):
                reject("employee", hn, "a place name in the FOREMAN NAME column",
                       "reject", sheet="TE", row=r, project=norm(proj))
                continue
            people.add(hn, "TE", job=jobnum)
            reject("assignment", hn, "foreman with a project but NO trailer id — "
                   "cannot create a custody row without the equipment it is for",
                   "warn", project=pname or norm(proj), job=jobnum or norm(job), sheet="TE", row=r)
        continue

    unit = f"TE-{int(mt.group(1)):03d}"
    hn = norm(holder)

    if unit in trailers:
        # TE repeats several ids as blank rows further down; a blank must never
        # overwrite a populated assignment.
        if not hn and not proj and not job:
            te_seen_blank_repeat.append((unit, r))
            continue

    status = None
    foreman = None
    if hn and STATUS_WORDS.search(hn):
        status = hn
    elif hn:
        foreman = hn
        people.add(hn, "TE", job=jobnum)

    rec = trailers.get(unit) or {
        "unit": unit, "foreman": None, "project": None, "job": None,
        "status": None, "source": "TE", "row": r,
    }
    rec["foreman"] = rec["foreman"] or foreman
    rec["project"] = rec["project"] or pname
    rec["job"] = rec["job"] or jobnum
    rec["status"] = rec["status"] or status
    trailers[unit] = rec

    if jobproblem:
        reject("project", pname or norm(proj) or unit, jobproblem, "warn",
               unit=unit, sheet="TE", row=r)

# ---------------------------------------------------------------- read TE-*
sheet_names = [n for n in wb.sheetnames if n.upper().startswith("TE-")]
tools = []
tool_seq = 0
sheet_meta = {}

for name in sheet_names:
    ws = wb[name]
    # header row: the one carrying DATE / DESCRIPTION
    header_row = None
    for r in range(1, min(ws.max_row, 12) + 1):
        vals = [norm(ws.cell(r, c).value).upper() for c in range(1, ws.max_column + 1)]
        if "DESCRIPTION" in vals or "DATE" in vals:
            header_row = r
            break
    if not header_row:
        reject("sheet", name, "no DATE/DESCRIPTION header row found; not parsed", "reject")
        continue

    # meta rows above the header carry the trailer id, the foreman and 'Project #job'
    unit = foreman = pname_raw = job = None
    for r in range(1, header_row):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = norm(v)
            m = TRAILER_RE.match(s)
            if m and not unit:
                unit = f"TE-{int(m.group(1)):03d}"
                continue
            hm = re.search(r"#\s*(\d{4,5})", s)
            if hm:
                job = hm.group(1)
                before = s[:hm.start()].strip(" #,;:-")
                if before and not foreman:
                    foreman = norm(before)
                continue
            if s and not foreman and not JOB_NUM.match(s) and not TRAILER_RE.match(s):
                foreman = s
    sheet_meta[name] = {"unit": unit, "foreman": foreman, "job": job}

    if not unit:
        reject("vehicle", name, "TE-* sheet with no readable trailer id in its header",
               "reject", sheet=name)

    # reconcile the sheet against TE
    if unit and unit in trailers:
        te_f = trailers[unit]["foreman"]
        if te_f and foreman and nkey(te_f) != nkey(foreman):
            reject("employee", unit,
                   f"foreman disagrees between sources: TE says {te_f!r}, sheet {name} says {foreman!r}",
                   "warn", unit=unit, sheet=name)
        if foreman:
            people.add(foreman, f"sheet:{name}", job=job)
        te_j = trailers[unit]["job"]
        if te_j and job and te_j != job:
            reject("project", unit,
                   f"job number disagrees between sources: TE says {te_j}, sheet {name} says {job}",
                   "warn", unit=unit, sheet=name)
    elif unit:
        if foreman:
            people.add(foreman, f"sheet:{name}", job=job)
        trailers[unit] = {
            "unit": unit, "foreman": foreman, "project": canon_project(None, job),
            "job": job, "status": None, "source": f"sheet:{name}", "row": None,
        }
        reject("vehicle", unit, "trailer exists only in its TE-* sheet, with no row in TE",
               "warn", sheet=name, foreman=foreman)

    # tools
    hdr = {}
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(header_row, c).value).upper()
        if h:
            hdr[h] = c

    def cell(r, label):
        c = hdr.get(label)
        return ws.cell(r, c).value if c else None

    for r in range(header_row + 1, ws.max_row + 1):
        desc = cell(r, "DESCRIPTION")
        make = cell(r, "MAKE")
        model = cell(r, "MODEL")
        serial = cell(r, "SERIAL #")
        other = cell(r, "OTHER")
        qty = cell(r, "QTY")
        if all(v is None for v in (desc, make, model, serial, qty)):
            continue
        d = norm(desc)
        if not d and not norm(make) and not norm(model):
            # a bare serial continuing the row above
            if serial is not None and tools:
                tools[-1]["extra_serials"].append(norm(serial))
            continue

        q = int(qty) if str(qty or "").strip().isdigit() else 1
        serials = [s for s in re.split(r"[/,]", norm(serial)) if s.strip()] if serial is not None else []
        tool_seq += 1
        tag = f"TOOL-{tool_seq:04d}"

        rec = {
            "tag": tag, "qty": q, "description": d or None,
            "make": norm(make) or None, "model": norm(model) or None,
            "serials": [s.strip() for s in serials], "extra_serials": [],
            "note": norm(other) or None,
            "trailer_unit": unit, "foreman": foreman,
            "job": job, "sheet": name,
        }
        tools.append(rec)

        if not d:
            reject("tool", tag, "no description; cannot identify the tool", "reject",
                   sheet=name, make=rec["make"], model=rec["model"])
        if q > 1 and len(serials) and len(serials) != q:
            reject("tool", tag,
                   f"qty {q} but {len(serials)} serial(s) — cannot tell how many "
                   "serialized units this row is",
                   "warn", sheet=name, description=d)

# blank repeats
if te_seen_blank_repeat:
    reject("vehicle", ",".join(sorted({u for u, _ in te_seen_blank_repeat})),
           f"{len(te_seen_blank_repeat)} blank repeat row(s) in TE discarded; a blank "
           "never overwrote a populated assignment",
           "info", sheet="TE")

# ---------------------------------------------------------------- trucks
truck_rows = []
seen_plate = {}
seen_vin = {}

for unit, holder, position, job, model, vin, plate, toll, year in TRUCKS:
    hn = norm(holder)
    jobnum, jobproblem = clean_job(job)

    if NOT_A_PERSON.search(hn):
        if re.search(r"out of service|remove it from portal", hn, re.I):
            reject("vehicle", unit, "source says out of service / remove from portal — "
                   "deliberately NOT imported", "reject",
                   holder=hn, vin=vin, license=plate)
            continue
        reject("vehicle", unit, f"holder is a department, not a person ({hn!r}) — "
               "truck is real but has no custodian to assign it to",
               "warn", vin=vin, license=plate, job=job)
        custodian = None
    else:
        custodian = hn
        people.add(hn, "truck-pdf", position=norm(position), job=jobnum)

    if not TRUCK_RE.match(unit) and not unit.upper().startswith("SUV"):
        reject("vehicle", unit, "unit number is not TRK-* or SUV-*", "warn")
    if unit.upper().startswith("SUV") and not re.match(r"^SUV-\d{3}$", unit):
        reject("vehicle", unit, f"unit number not zero-padded like its siblings "
               f"({unit} vs SUV-001) — will not sort or match consistently", "warn")

    if plate:
        p = plate.upper()
        if p in seen_plate:
            reject("vehicle", unit, f"duplicate license plate {plate!r}, already on "
                   f"{seen_plate[p]} — one of the two is wrong", "reject",
                   vin=vin, holder=hn)
        else:
            seen_plate[p] = unit
        if not re.match(r"^[A-Z0-9-]{5,9}$", p):
            reject("vehicle", unit, f"license is not a plate: {plate!r}", "reject",
                   vin=vin, holder=hn)
    else:
        reject("vehicle", unit, "no license plate", "warn", vin=vin, holder=hn)

    if vin:
        if vin in seen_vin:
            reject("vehicle", unit, f"duplicate VIN {vin!r}, already on {seen_vin[vin]}",
                   "reject", holder=hn)
        else:
            seen_vin[vin] = unit
        if not VALID_VIN.match(vin):
            reject("vehicle", unit, f"VIN is not 17 valid characters: {vin!r}", "reject",
                   holder=hn)
    else:
        reject("vehicle", unit, "no VIN", "warn", holder=hn)

    if year and int(year) > 2025:
        reject("vehicle", unit, f"model year {year} is in the future", "warn",
               vin=vin, holder=hn)

    if jobproblem:
        reject("vehicle", unit, f"job cell unusable — {jobproblem}", "warn",
               holder=hn, note="truck can load; its project link cannot")

    truck_rows.append({
        "unit": unit, "code": unit,
        "equipment_kind": "vehicle", "vehicle_type": "truck",
        "make_model": norm(model), "plate": plate, "vin": vin,
        "year": year, "toll_tag": "yes" if (toll or "").lower() == "yes" else "no",
        "custodian_name": canon_name(custodian), "position": norm(position),
        "job": jobnum, "job_raw": norm(job) if job else None,
        "ownership": "company_owned",
    })

# ------------------------------------------------- one truck per foreman
# `vehicle_one_truck_per_foreman_uq` is a REAL unique index on
# (tenant_id, foreman_employee_id) WHERE vehicle_type='truck' AND
# ownership_type='company_owned' (migration 0030). A second company truck for
# the same person does not warn, it aborts the load. So the first truck keeps
# the custodian and any later one is loaded with none, rather than dropping a
# real vehicle out of the register.
#
# Note the index is deliberately TRUCKS ONLY -- see the rationale at
# schema/location.ts:161, which records that one Urban foreman really does run
# two loaded trailers. Do not generalise this to trailers.
_truck_held = {}
for t in truck_rows:
    if t["vehicle_type"] != "truck" or t["ownership"] != "company_owned":
        continue
    who = t["custodian_name"]
    if not who:
        continue
    k = nkey(who)
    if k in _truck_held:
        reject("vehicle", t["unit"],
               f"{who} already holds {_truck_held[k]}, and the register allows one "
               "company truck per person (vehicle_one_truck_per_foreman_uq). This "
               "truck loads with NO custodian — confirm which vehicle is really theirs",
               "warn", vin=t["vin"], license=t["plate"])
        t["custodian_name"] = None
    else:
        _truck_held[k] = t["unit"]

# ---------------------------------------------------------------- projects
projects = OrderedDict()


def add_project(name, job):
    if not name and not job:
        return None
    key = job or nkey(name)
    if key not in projects:
        projects[key] = {"external_id": job, "name": name or f"Job {job}", "status": "in_progress"}
    elif name and projects[key]["name"].startswith("Job "):
        projects[key]["name"] = name
    return key


# The yard first, so everything unresolvable has somewhere to land.
add_project(YARD, None)
for t in trailers.values():
    add_project(t["project"], t["job"])
for t in truck_rows:
    if t["job"]:
        add_project(JOB_CANON.get(t["job"]), t["job"])
for t in tools:
    if t["job"]:
        add_project(JOB_CANON.get(t["job"]), t["job"])

# DECISIONS 3: the yard is the yard and carries no job number; 24002 is a real
# job of its own. Enforce both rather than trusting whichever row was seen first.
yard_key = nkey(YARD)
if yard_key in projects:
    projects[yard_key]["external_id"] = None
    projects[yard_key]["name"] = YARD

if "24002" in projects:
    p = projects["24002"]
    if nkey(p["name"]) == yard_key or p["name"].startswith("Job "):
        # TE-007 typed the yard against 24002. The truck list proves 24002 is a
        # real job, but names nobody's project — so the name is genuinely unknown.
        p["name"] = "Job 24002"
        reject("project", "24002",
               "resolved as a real job, separate from the Equipment Yard (it has a "
               "superintendent TRK-036, surveyor TRK-038, traffic control foreman "
               "TRK-020 and field engineer TRK-045). Its PROPER NAME is not in "
               "either source — loaded as 'Job 24002' and needs renaming",
               "warn", crew="TRK-020, TRK-036, TRK-038, TRK-045")

# TE-007 named the yard but carried job 24002; the trailer keeps the yard.
for t in trailers.values():
    if t["job"] == "24002" and t["project"] and nkey(t["project"]) == yard_key:
        t["job"] = None
        reject("vehicle", t["unit"],
               "TE lists this trailer as YARD with job #24002; the yard carries no "
               "job number, so the job link is dropped and the trailer stays at the "
               "Equipment Yard", "info")

# ---------------------------------------------------------------- ambiguous people
settled = {frozenset((nkey(a), nkey(b))) for a, b in KEPT_SEPARATE}
for a, b, why in people.ambiguities():
    pair = frozenset((nkey(a["name"]), nkey(b["name"])))
    if pair in settled:
        reject("employee", f"{a['name']} | {b['name']}",
               f"near-match reviewed and DELIBERATELY kept separate ({why}). "
               "Loading both as two employees is the intended outcome; merge in "
               "the app later if they turn out to be one person",
               "info",
               a_jobs=", ".join(a["jobs"]) or None, b_jobs=", ".join(b["jobs"]) or None)
        continue
    reject("employee", f"{a['name']} | {b['name']}",
           f"possible duplicate person — {why}. NOT merged, and NOT yet ruled on. "
           f"{a['name']}: jobs={a['jobs'] or '-'} sources={a['sources']}; "
           f"{b['name']}: jobs={b['jobs'] or '-'} sources={b['sources']}",
           "reject")

# people holding more than one truck / appearing on several jobs
holders = defaultdict(list)
for t in truck_rows:
    if t["custodian_name"]:
        holders[pkey(t["custodian_name"])].append(t["unit"])
for k, units in holders.items():
    if len(units) > 1:
        reject("assignment", people.by_name[k]["name"],
               f"holds {len(units)} vehicles ({', '.join(units)}) — legal, but confirm "
               "it is one person and not two people sharing a name",
               "warn")

# ---------------------------------------------------------------- trailers out
# One person holding two enclosed trailers is not how Urban runs, so it is a
# data question every time rather than something to load quietly. TE-027 is the
# known case and is resolved in DECISIONS 4; a NEW one must stop the load.
held = defaultdict(list)
for t in trailers.values():
    if t["foreman"] and t["unit"] not in NO_CUSTODIAN_TRAILERS:
        held[pkey(t["foreman"])].append(t["unit"])
for k, units in held.items():
    if len(units) > 1:
        reject("vehicle", ", ".join(sorted(units)),
               f"{people.by_name[k]['name']!r} holds {len(units)} enclosed trailers — "
               "one person runs one trailer, so either these are two people with the "
               "same name or one row is stale. Not resolvable from the source",
               "reject", custodian=people.by_name[k]["name"])

trailer_rows = []
for t in trailers.values():
    unit = t["unit"]
    proj = t["project"]
    job = t["job"]
    custodian = t["foreman"]

    # DECISIONS 4: loaded with no custodian on purpose.
    if unit in NO_CUSTODIAN_TRAILERS:
        if custodian:
            reject("vehicle", unit,
                   f"source names {custodian!r} as custodian, but this trailer is "
                   "loaded with NO custodian by decision — that person already holds "
                   "another trailer. Its tools stay with the trailer, assigned to "
                   "nobody, until Urban says who runs it",
                   "info", project=proj, job=job)
        custodian = None

    if not custodian and unit not in NO_CUSTODIAN_TRAILERS:
        why = (f"flagged in source as {t['status']!r}" if t["status"]
               else "no foreman in TE and none in a TE-* sheet")
        reject("vehicle", unit, f"trailer has no custodian — {why}", "warn",
               project=proj, job=job)
    if not proj and not job:
        # "anything confusing, especially on smalltools, goes to Equipment Yard"
        proj = YARD
        reject("vehicle", unit, "no project or job anywhere — parked at Equipment Yard",
               "info", custodian=custodian)

    trailer_rows.append({
        "unit": unit, "code": unit,
        "equipment_kind": "attachment", "vehicle_type": "trailer",
        "make_model": "Enclosed trailer", "plate": None, "vin": None,
        "year": None, "toll_tag": "no",
        "custodian_name": canon_name(custodian), "position": "Foreman" if custodian else None,
        "job": job, "job_raw": job, "ownership": "company_owned",
    })

# ---------------------------------------------------------------- tools out
tool_rows = []
# Where a trailer ended up, after the custodian decisions above. A tool follows
# the TRAILER it is aboard — its project and its custodian are the trailer's, not
# the sheet header's, so a trailer loaded with no custodian carries unassigned
# tools rather than sending them to the yard (DECISIONS 4).
trailer_by_unit = {r["unit"]: r for r in trailer_rows}

for t in tools:
    unit = t["trailer_unit"]
    veh = trailer_by_unit.get(unit) if unit else None

    if veh:
        custodian = veh["custodian_name"]
        proj_job = veh["job"]
        project_name = JOB_CANON.get(proj_job) if proj_job else None
        if not project_name and not proj_job:
            project_name = YARD
    else:
        # No trailer to sit on: "anything confusing, especially on small tools,
        # goes to Equipment Yard."
        custodian = None
        proj_job = None
        project_name = YARD
        reject("tool", t["tag"],
               f"no trailer resolved from sheet {t['sheet']!r} — parked at Equipment Yard",
               "info", description=t["description"])

    serials = t["serials"] + t["extra_serials"]
    tool_rows.append({
        "tag": t["tag"],
        "description": t["description"],
        "make": t["make"], "model": t["model"],
        "serial": serials[0] if serials else None,
        "extra_serials": ";".join(serials[1:]) if len(serials) > 1 else None,
        "qty": t["qty"],
        "trailer_unit": unit,
        "custodian_name": canon_name(custodian),
        "job": proj_job,
        "project_name": project_name,
        # serial -> Equipment Department, none -> Purchased Department
        "department": "Equipment Department" if serials else "Purchased Department",
        "note": t["note"],
        "source_sheet": t["sheet"],
    })

# ---------------------------------------------------------------- employees out
emp_rows = []
for k, rec in people.by_name.items():
    emp_rows.append({
        "name": rec["name"],
        "positions": " | ".join(rec["positions"]) or None,
        "jobs": " | ".join(rec["jobs"]) or None,
        "sources": " | ".join(rec["sources"]),
    })

# ---------------------------------------------------------------- write
import os
os.makedirs(OUT, exist_ok=True)


def write_csv(path, rows, cols):
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


write_csv(f"{OUT}/projects.csv", list(projects.values()),
          ["external_id", "name", "status"])
write_csv(f"{OUT}/employees.csv", emp_rows,
          ["name", "positions", "jobs", "sources"])
write_csv(f"{OUT}/vehicles.csv", truck_rows + trailer_rows,
          ["unit", "code", "equipment_kind", "vehicle_type", "make_model", "plate",
           "vin", "year", "toll_tag", "custodian_name", "position", "job", "job_raw",
           "ownership"])
write_csv(f"{OUT}/tools.csv", tool_rows,
          ["tag", "description", "make", "model", "serial", "extra_serials", "qty",
           "trailer_unit", "custodian_name", "job", "project_name", "department",
           "note", "source_sheet"])

sev_order = {"reject": 0, "warn": 1, "info": 2}
rejects.sort(key=lambda r: (sev_order.get(r["severity"], 9), r["entity"], str(r["id"])))

with open(f"{OUT}/rejects.json", "w") as f:
    json.dump({
        "purpose": "Rows that are NOT safe to load, and rows that load but carry a "
                   "question. severity=reject means loading it would collide, "
                   "misidentify a person, or invent an id. Every one needs a human "
                   "decision; none is auto-resolved.",
        "source_files": ["docs/data/TOOL LIST BY NAME NEW.xlsx",
                         "Latest update on company truck.pdf (transcribed into build_import.py)"],
        "counts": {
            "reject": sum(1 for r in rejects if r["severity"] == "reject"),
            "warn": sum(1 for r in rejects if r["severity"] == "warn"),
            "info": sum(1 for r in rejects if r["severity"] == "info"),
        },
        "rows": rejects,
    }, f, indent=2, ensure_ascii=False)

with open(f"{OUT}/rejects.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["severity", "entity", "id", "reason", "raw"])
    for r in rejects:
        w.writerow([r["severity"], r["entity"], r["id"], r["reason"],
                    json.dumps(r["raw"], ensure_ascii=False)])

print(f"projects       {len(projects)}")
print(f"employees      {len(emp_rows)}")
print(f"trucks         {len(truck_rows)}")
print(f"trailers       {len(trailer_rows)}")
print(f"tools          {len(tool_rows)}")
print(f"rejects        reject={sum(1 for r in rejects if r['severity']=='reject')} "
      f"warn={sum(1 for r in rejects if r['severity']=='warn')} "
      f"info={sum(1 for r in rejects if r['severity']=='info')}")
print(f"\nwrote -> {OUT}/")
